from __future__ import annotations

from dataclasses import dataclass, field
import json
import re
import struct
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

from ..core.config import BASE_DIR
from ..core.conversions import (
    BADGE_LEVEL_NAMES,
    HEIGHT_MAX_INCHES,
    HEIGHT_MIN_INCHES,
    convert_raw_to_minmax_potential,
    convert_raw_to_rating,
    convert_tendency_raw_to_rating,
    raw_height_to_inches,
)
from ..core import offsets as offsets_mod
from ..core.offsets import NAME_MAX_CHARS, PLAYER_STRIDE
from ..core.perf import timed
from ..models.data_model import PlayerDataModel
from ..models.player import Player

try:  # Lazy dependency
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - handled at runtime
    openpyxl = None


_ENTITY_CONFIG: dict[str, dict[str, Any]] = {
    "players": {
        "label": "Players",
        "super_type": "Players",
        "key_sheets": ["Vitals"],
        "name_tokens": {
            "first": {"First Name"},
            "last": {"Last Name"},
            "full": {"Name", "Player Name", "Full Name"},
        },
    },
    "teams": {
        "label": "Teams",
        "super_type": "Teams",
        "key_sheets": ["Team Vitals", "Teams", "Team Info"],
        "name_tokens": {"full": {"Team Name", "Team Vitals - TEAMNAME"}},
    },
    "staff": {
        "label": "Staff",
        "super_type": "Staff",
        "key_sheets": ["Staff Vitals"],
        "name_tokens": {
            "first": {"Staff Vitals - FIRSTNAME"},
            "last": {"Staff Vitals - LASTNAME"},
            "full": {"Staff Vitals - NAME"},
        },
    },
    "stadiums": {
        "label": "Stadiums",
        "super_type": "Stadiums",
        "key_sheets": ["Stadium"],
        "name_tokens": {"full": {"Stadium Vitals - NAME", "Arena Name"}},
    },
}

_FALLBACK = object()


@dataclass(frozen=True)
class _RecordSnapshot:
    base_addr: int
    stride: int
    buffer: memoryview
    max_count: int

    def record_view(self, index: int) -> memoryview | None:
        if index < 0 or index >= self.max_count:
            return None
        start = index * self.stride
        end = start + self.stride
        if end > len(self.buffer):
            return None
        return self.buffer[start:end]


def _build_player_snapshot(
    model: PlayerDataModel,
    entities: list[tuple[int, int | None]],
) -> _RecordSnapshot | None:
    with timed("excel_import.build_player_snapshot"):
        if not entities:
            return None
        if PLAYER_STRIDE <= 0:
            return None
        if not model.mem.open_process():
            return None
        base = model._resolve_player_base_ptr()
        if base is None:
            return None
        max_index = max(idx for idx, _ in entities)
        max_count = min(max_index + 1, model.max_players)
        team_base = model._resolve_team_base_ptr()
        if team_base is not None and team_base > base:
            max_before_team = int((team_base - base) // PLAYER_STRIDE)
            if max_before_team > 0:
                max_count = min(max_count, max_before_team)
        if max_count <= 0:
            return None
        total_bytes = max_count * PLAYER_STRIDE
        try:
            blob = model.mem.read_bytes(base, total_bytes)
        except Exception:
            return None
        return _RecordSnapshot(base_addr=base, stride=PLAYER_STRIDE, buffer=memoryview(blob), max_count=max_count)


def _decode_string_from_record(record: memoryview, offset: int, max_chars: int, encoding: str) -> object:
    if offset < 0 or max_chars <= 0:
        return _FALLBACK
    if encoding == "ascii":
        byte_len = max_chars
    else:
        byte_len = max_chars * 2
    end = offset + byte_len
    if end > len(record):
        return _FALLBACK
    raw = record[offset:end].tobytes()
    try:
        if encoding == "ascii":
            text = raw.decode("ascii", errors="ignore")
        else:
            text = raw.decode("utf-16le", errors="ignore")
    except Exception:
        return _FALLBACK
    zero = text.find("\x00")
    if zero != -1:
        text = text[:zero]
    return text


def _decode_float_from_record(record: memoryview, offset: int, byte_len: int) -> object:
    need = 8 if byte_len >= 8 else 4
    end = offset + need
    if offset < 0 or end > len(record):
        return _FALLBACK
    raw = record[offset:end].tobytes()
    try:
        return struct.unpack("<d" if byte_len >= 8 else "<f", raw)[0]
    except Exception:
        return _FALLBACK


def _decode_bits_from_record(record: memoryview, offset: int, start_bit: int, length_bits: int) -> object:
    if length_bits <= 0:
        return _FALLBACK
    bits_needed = start_bit + length_bits
    bytes_needed = (bits_needed + 7) // 8
    end = offset + bytes_needed
    if offset < 0 or end > len(record):
        return _FALLBACK
    raw = int.from_bytes(record[offset:end], "little")
    raw >>= start_bit
    mask = (1 << length_bits) - 1
    return raw & mask


def _decode_field_value_from_record(
    model: PlayerDataModel,
    *,
    entity_type: str,
    category: str,
    field_name: str,
    meta: dict[str, object],
    record: memoryview | None,
) -> object:
    if record is None:
        return _FALLBACK
    (
        offset,
        start_bit,
        length_bits,
        requires_deref,
        deref_offset,
        field_type,
        byte_length,
        values,
    ) = model._extract_field_parts(meta)
    if requires_deref or deref_offset:
        return _FALLBACK
    field_type_norm = model._normalize_field_type(field_type)
    length_raw = length_bits
    if length_bits <= 0 and byte_length > 0:
        length_bits = byte_length * 8
    name_lower = str(field_name or "").strip().lower()
    category_lower = str(category or "").strip().lower()
    if model._is_string_type(field_type_norm):
        max_chars = length_raw if length_raw > 0 else byte_length
        if max_chars <= 0:
            if "name" in name_lower and NAME_MAX_CHARS > 0:
                max_chars = NAME_MAX_CHARS
            else:
                max_chars = 64
        enc = model._string_encoding_for_type(field_type_norm)
        return _decode_string_from_record(record, offset, max_chars, enc)
    if entity_type.strip().lower() == "player" and name_lower == "weight":
        val = _decode_float_from_record(record, offset, 4)
        if val is _FALLBACK:
            return _FALLBACK
        if not isinstance(val, (int, float)):
            return _FALLBACK
        return int(round(val))
    if model._is_float_type(field_type_norm):
        byte_len = model._effective_byte_length(byte_length, length_bits, default=4)
        return _decode_float_from_record(record, offset, byte_len)
    raw_val = _decode_bits_from_record(record, offset, start_bit, length_bits)
    if raw_val is _FALLBACK:
        return _FALLBACK
    if not isinstance(raw_val, int):
        return _FALLBACK
    raw_int = raw_val
    if values:
        idx = model._clamp_enum_index(raw_int, values, length_bits)
        return idx
    if model._is_pointer_type(field_type_norm) or model._is_color_type(field_type_norm):
        if model._is_team_pointer_field(entity_type, category, field_name, field_type_norm):
            team_name = model._team_pointer_to_display_name(raw_int)
            if team_name:
                return team_name
        return model._format_hex_value(raw_int, length_bits, byte_length)
    if entity_type.strip().lower() == "player" and name_lower == "height":
        inches = raw_height_to_inches(raw_int)
        if inches < HEIGHT_MIN_INCHES:
            inches = HEIGHT_MIN_INCHES
        if inches > HEIGHT_MAX_INCHES:
            inches = HEIGHT_MAX_INCHES
        return inches
    if category_lower in ("attributes", "durability"):
        return convert_raw_to_rating(raw_int, length_bits or 8)
    if category_lower == "potential":
        if "min" in name_lower or "max" in name_lower:
            return convert_raw_to_minmax_potential(raw_int, length_bits or 8)
        return convert_raw_to_rating(raw_int, length_bits or 8)
    if category_lower == "tendencies":
        return convert_tendency_raw_to_rating(raw_int, length_bits or 8)
    if category_lower == "badges":
        max_lvl = max(0, len(BADGE_LEVEL_NAMES) - 1)
        if raw_int < 0:
            return 0
        if raw_int > max_lvl:
            return max_lvl
        return raw_int
    return raw_int


@dataclass(frozen=True)
class FieldMatch:
    category: str
    field_name: str
    meta: dict[str, object]


@dataclass
class ImportResult:
    entity_type: str
    workbook: Path
    rows_seen: int = 0
    rows_applied: int = 0
    fields_written: int = 0
    missing_names: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    skipped_sheets: list[str] = field(default_factory=list)
    ignored_columns: dict[str, int] = field(default_factory=dict)

    def summary_text(self) -> str:
        label = _ENTITY_CONFIG.get(self.entity_type, {}).get("label", self.entity_type.title())
        lines = [
            f"{label} import: {self.rows_applied}/{self.rows_seen} rows updated",
            f"Fields written: {self.fields_written}",
        ]
        if self.skipped_sheets:
            lines.append(f"Sheets skipped: {', '.join(self.skipped_sheets)}")
        if self.ignored_columns:
            total = sum(self.ignored_columns.values())
            lines.append(f"Columns ignored: {total}")
        if self.missing_names:
            lines.append(f"Missing names: {len(self.missing_names)}")
        if self.warnings:
            lines.append(f"Warnings: {len(self.warnings)}")
        if self.errors:
            lines.append(f"Errors: {len(self.errors)}")
        return "\n".join(lines)


@dataclass
class ExportResult:
    entity_type: str
    workbook: Path
    rows_written: int = 0
    sheets_written: int = 0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def summary_text(self) -> str:
        label = _ENTITY_CONFIG.get(self.entity_type, {}).get("label", self.entity_type.title())
        lines = [
            f"{label} export: {self.rows_written} rows written",
            f"Sheets written: {self.sheets_written}",
        ]
        if self.warnings:
            lines.append(f"Warnings: {len(self.warnings)}")
        if self.errors:
            lines.append(f"Errors: {len(self.errors)}")
        return "\n".join(lines)


def template_path_for(entity_type: str) -> Path:
    key = (entity_type or "").strip().lower()
    name_map = {
        "players": "ImportPlayers.xlsx",
        "teams": "ImportTeams.xlsx",
        "staff": "ImportStaff.xlsx",
        "stadiums": "ImportStadiums.xlsx",
        "stadium": "ImportStadiums.xlsx",
    }
    filename = name_map.get(key)
    if not filename:
        raise ValueError(f"Unknown entity type: {entity_type}")
    return BASE_DIR / "Offsets" / filename


def _ensure_openpyxl() -> Any:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for Excel import/export.")
    return openpyxl


def _sanitize_excel_value(value: object) -> Any:
    if not isinstance(value, str):
        return value
    if not value:
        return value
    try:
        from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE  # type: ignore
    except Exception:
        ILLEGAL_CHARACTERS_RE = None
    if ILLEGAL_CHARACTERS_RE is not None:
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    cleaned = []
    for ch in value:
        code = ord(ch)
        if code in (9, 10, 13):
            cleaned.append(ch)
            continue
        if 0 <= code <= 8 or 11 <= code <= 12 or 14 <= code <= 31:
            continue
        cleaned.append(ch)
    return "".join(cleaned)


def _header_text(text: object) -> str:
    return str(text or "").strip()


def _field_text(text: object) -> str:
    return str(text or "").strip()


def _header_candidates(header: object) -> list[str]:
    raw = _header_text(header)
    if not raw:
        return []
    lower = raw.lower()
    if lower != raw:
        return [raw, lower]
    return [raw]


def _preferred_categories(sheet_name: str, categories: dict[str, list[dict]]) -> list[str]:
    sheet_lower = (sheet_name or "").strip().lower()
    if not sheet_lower:
        return []
    direct = [cat for cat in categories if cat.lower() == sheet_lower]
    if direct:
        return direct
    fuzzy: list[str] = []
    for cat in categories:
        cat_lower = cat.lower()
        if sheet_lower in cat_lower or cat_lower in sheet_lower:
            fuzzy.append(cat)
    return fuzzy


def _build_field_lookup(
    model: PlayerDataModel,
    categories: dict[str, list[dict]],
) -> tuple[dict[str, list[FieldMatch]], dict[str, dict[str, FieldMatch]]]:
    global_map: dict[str, list[FieldMatch]] = {}
    per_category: dict[str, dict[str, FieldMatch]] = {}
    by_category_name: dict[tuple[str, str], FieldMatch] = {}
    for cat_name, fields in categories.items():
        cat_map: dict[str, FieldMatch] = {}
        for field in fields:
            if not isinstance(field, dict):
                continue
            name = str(field.get("name") or "").strip()
            if not name:
                continue
            key = _field_text(name)
            if not key:
                continue
            key_lower = key.lower()
            category_label = str(field.get("category") or cat_name)
            match = FieldMatch(category_label, name, field)
            cat_map.setdefault(key, match)
            global_map.setdefault(key, []).append(match)
            if key_lower != key:
                cat_map.setdefault(key_lower, match)
                global_map.setdefault(key_lower, []).append(match)
            by_category_name[(category_label, name)] = match
        per_category[cat_name] = cat_map
    _augment_with_display_aliases(global_map, per_category, by_category_name)
    return global_map, per_category


def _resolve_version_key(raw: dict) -> str | None:
    versions = raw.get("versions")
    if not isinstance(versions, dict) or not versions:
        return None
    target = getattr(offsets_mod, "_current_offset_target", None) or offsets_mod.MODULE_NAME
    version_hint = None
    if isinstance(target, str):
        match = re.search(r"2k(\d{2})", target.lower())
        if match:
            version_hint = f"2k{match.group(1)}"
    if version_hint:
        for key in versions.keys():
            if version_hint in str(key).lower():
                return str(key)
    return str(next(iter(versions.keys())))


def _augment_with_display_aliases(
    global_map: dict[str, list[FieldMatch]],
    per_category: dict[str, dict[str, FieldMatch]],
    by_category_name: dict[tuple[str, str], FieldMatch],
) -> None:
    offsets_path = BASE_DIR / "Offsets" / "offsets.json"
    if not offsets_path.is_file():
        return
    try:
        with offsets_path.open("r", encoding="utf-8") as handle:
            raw = json.load(handle)
    except Exception:
        return
    if not isinstance(raw, dict):
        return
    offsets_list = raw.get("offsets")
    if not isinstance(offsets_list, list):
        return
    version_key = _resolve_version_key(raw)
    if not version_key:
        return
    for entry in offsets_list:
        if not isinstance(entry, dict):
            continue
        display = entry.get("display_name")
        variant_names = entry.get("variant_names")
        per_version = entry.get("versions")
        if not isinstance(per_version, dict):
            continue
        v_entry = per_version.get(version_key)
        if not isinstance(v_entry, dict):
            continue
        category = v_entry.get("category") or entry.get("canonical_category") or entry.get("super_type") or ""
        name = (
            v_entry.get("name")
            or entry.get("display_name")
            or entry.get("normalized_name")
            or entry.get("canonical_name")
            or ""
        )
        cat_name = str(category).strip()
        field_name = str(name).strip()
        if not cat_name or not field_name:
            continue
        match = by_category_name.get((cat_name, field_name))
        if not match:
            continue
        aliases: set[str] = set()
        if display:
            aliases.add(str(display))
        if isinstance(variant_names, (list, tuple)):
            for name in variant_names:
                if name:
                    aliases.add(str(name))
        for alias_raw in aliases:
            alias = _field_text(alias_raw)
            if not alias:
                continue
            lower = alias.lower()
            global_map.setdefault(alias, []).append(match)
            per_category.setdefault(match.category, {}).setdefault(alias, match)
            if lower != alias:
                global_map.setdefault(lower, []).append(match)
                per_category.setdefault(match.category, {}).setdefault(lower, match)


def _map_headers(
    model: PlayerDataModel,
    headers: Sequence[object],
    global_map: dict[str, list[FieldMatch]],
    per_category: dict[str, dict[str, FieldMatch]],
    preferred: Sequence[str],
) -> tuple[list[FieldMatch | None], list[str]]:
    mapped: list[FieldMatch | None] = []
    ignored: list[str] = []
    for header in headers:
        if header is None or str(header).strip() == "":
            mapped.append(None)
            continue
        candidates = _header_candidates(header)
        match: FieldMatch | None = None
        for key in candidates:
            for cat in preferred:
                match = per_category.get(cat, {}).get(key)
                if match:
                    break
            if match:
                break
        if match is None:
            ignored.append(str(header))
        mapped.append(match)
    return mapped, ignored


def _find_column(headers: Sequence[object], candidates: set[str]) -> int | None:
    for idx, header in enumerate(headers):
        for key in _header_candidates(header):
            if key in candidates:
                return idx
    return None


def _resolve_row_name(
    headers: Sequence[object],
    row: Sequence[object],
    model: PlayerDataModel,
    name_tokens: dict[str, set[str]],
    row_key_override: str | None,
) -> str | None:
    if row_key_override:
        return row_key_override
    values = list(row)
    first_idx = _find_column(headers, name_tokens.get("first", set()))
    last_idx = _find_column(headers, name_tokens.get("last", set()))
    full_idx = _find_column(headers, name_tokens.get("full", set()))
    if full_idx is not None:
        raw = values[full_idx]
        name = str(raw or "").strip()
        if name:
            return name
    first = str(values[first_idx] or "").strip() if first_idx is not None else ""
    last = str(values[last_idx] or "").strip() if last_idx is not None else ""
    combined = f"{first} {last}".strip()
    return combined or None


def _build_row_key_map(
    workbook,
    model: PlayerDataModel,
    config: dict[str, Any],
) -> dict[int, str]:
    row_key_map: dict[int, str] = {}
    key_sheets = [s.lower() for s in config.get("key_sheets", [])]
    if not key_sheets:
        return row_key_map
    target_sheet = None
    for name in workbook.sheetnames:
        if name.strip().lower() in key_sheets:
            target_sheet = workbook[name]
            break
    if target_sheet is None:
        return row_key_map
    headers = [cell.value for cell in next(target_sheet.iter_rows(min_row=1, max_row=1))]
    tokens = config.get("name_tokens", {})
    for idx, row in enumerate(target_sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = _resolve_row_name(headers, row, model, tokens, None)
        if name:
            row_key_map[idx] = name
    return row_key_map


def _row_has_values(row: Iterable[object]) -> bool:
    for value in row:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def _resolve_player(model: PlayerDataModel, name: str, player_index_map: dict[int, Player]) -> tuple[int, int | None] | None:
    if not name:
        return None
    try:
        idx = int(str(name).strip(), 0)
        player = player_index_map.get(idx)
        if player is not None:
            return (player.index, player.record_ptr)
    except Exception:
        pass
    indices = model.find_player_indices_by_name(name)
    if not indices:
        return None
    player = player_index_map.get(indices[0])
    if player is None:
        return None
    return (player.index, player.record_ptr)


def _resolve_named_index(name: str, name_map: dict[str, int]) -> int | None:
    if not name:
        return None
    try:
        idx = int(str(name).strip(), 0)
        return idx
    except Exception:
        pass
    return name_map.get(name.strip().lower())


def import_excel_workbook(
    model: PlayerDataModel,
    workbook_path: str | Path,
    entity_type: str,
    *,
    name_overrides: dict[str, str] | None = None,
    only_names: set[str] | None = None,
    progress_cb: Callable[[int, int, str | None], None] | None = None,
) -> ImportResult:
    with timed("excel_import.import_workbook"):
        oxl = _ensure_openpyxl()
        entity_key = (entity_type or "").strip().lower()
        config = _ENTITY_CONFIG.get(entity_key)
        if not config:
            raise ValueError(f"Unknown entity type: {entity_type}")
        workbook = Path(workbook_path)
        result = ImportResult(entity_type=entity_key, workbook=workbook)
        categories = model.get_categories_for_super(config["super_type"])
        if not categories:
            result.warnings.append(f"No categories loaded for {config['super_type']}.")
            return result
        global_map, per_category = _build_field_lookup(model, categories)
        wb = oxl.load_workbook(workbook, data_only=True)
        row_key_map = _build_row_key_map(wb, model, config)
        player_index_map: dict[int, Player] = {p.index: p for p in model.players}
        team_name_map = {name.lower(): idx for idx, name in model.team_list}
        staff_name_map = {name.lower(): idx for idx, name in model.staff_list}
        stadium_name_map = {name.lower(): idx for idx, name in model.stadium_list}
        missing: list[str] = []
        missing_seen: set[str] = set()
        progress_total = 0
        progress_current = 0
        if progress_cb is not None:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_lower = sheet_name.strip().lower()
                if entity_key == "teams" and sheet_lower == "team players":
                    continue
                if sheet.max_row > 1:
                    progress_total += sheet.max_row - 1
            if progress_total > 0:
                progress_cb(0, progress_total, None)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_lower = sheet_name.strip().lower()
            if entity_key == "teams" and sheet_lower == "team players":
                result.skipped_sheets.append(sheet_name)
                continue
            header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
            if not header_row:
                continue
            headers = [cell.value for cell in header_row]
            preferred = _preferred_categories(sheet_name, categories)
            header_map, ignored_headers = _map_headers(model, headers, global_map, per_category, preferred)
            if ignored_headers:
                result.ignored_columns[sheet_name] = len(ignored_headers)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if progress_cb is not None and progress_total > 0:
                    progress_current += 1
                    if progress_current % 5 == 0 or progress_current == progress_total:
                        progress_cb(progress_current, progress_total, sheet_name)
                if not _row_has_values(row):
                    continue
                result.rows_seen += 1
                row_name_hint = row_key_map.get(row_idx)
                raw_name = _resolve_row_name(headers, row, model, config.get("name_tokens", {}), row_name_hint)
                if raw_name is None:
                    continue
                if only_names is not None and raw_name not in only_names:
                    continue
                mapped_name = raw_name
                if name_overrides and raw_name in name_overrides:
                    mapped_name = name_overrides.get(raw_name) or raw_name
                target = None
                record_ptr = None
                if entity_key == "players":
                    resolved = _resolve_player(model, mapped_name, player_index_map)
                    if resolved is not None:
                        target, record_ptr = resolved
                elif entity_key == "teams":
                    target = _resolve_named_index(mapped_name, team_name_map)
                elif entity_key == "staff":
                    target = _resolve_named_index(mapped_name, staff_name_map)
                elif entity_key == "stadiums":
                    target = _resolve_named_index(mapped_name, stadium_name_map)
                if target is None:
                    if raw_name not in missing_seen:
                        missing_seen.add(raw_name)
                        missing.append(raw_name)
                    continue
                row_applied = False
                for col_idx, value in enumerate(row):
                    match = header_map[col_idx] if col_idx < len(header_map) else None
                    if match is None:
                        continue
                    if value is None:
                        continue
                    if isinstance(value, str) and not value.strip():
                        continue
                    try:
                        ok = model.encode_field_value(
                            entity_type=entity_key[:-1] if entity_key.endswith("s") else entity_key,
                            entity_index=int(target),
                            category=match.category,
                            field_name=match.field_name,
                            meta=match.meta,
                            display_value=value,
                            record_ptr=record_ptr,
                        )
                    except Exception as exc:
                        result.errors.append(f"{sheet_name} row {row_idx}: {exc}")
                        ok = False
                    if ok:
                        row_applied = True
                        result.fields_written += 1
                if row_applied:
                    result.rows_applied += 1
        result.missing_names = missing
        return result


def export_excel_workbook(
    model: PlayerDataModel,
    output_path: str | Path,
    entity_type: str,
    *,
    template_path: str | Path | None = None,
    progress_cb: Callable[[int, int, str | None], None] | None = None,
    team_filter: set[str] | None = None,
    players: Sequence[Player] | None = None,
) -> ExportResult:
    with timed("excel_import.export_workbook"):
        oxl = _ensure_openpyxl()
        entity_key = (entity_type or "").strip().lower()
        config = _ENTITY_CONFIG.get(entity_key)
        if not config:
            raise ValueError(f"Unknown entity type: {entity_type}")
        template = Path(template_path) if template_path else template_path_for(entity_key)
        output = Path(output_path)
        result = ExportResult(entity_type=entity_key, workbook=output)
        categories = model.get_categories_for_super(config["super_type"])
        if not categories:
            result.warnings.append(f"No categories loaded for {config['super_type']}.")
        global_map, per_category = _build_field_lookup(model, categories)
        wb = oxl.load_workbook(template)
        sanitized_values = 0
        snapshot: _RecordSnapshot | None = None
        if entity_key == "players":
            if players is not None:
                filtered_players = list(players)
            else:
                filtered_players = list(model.players)
                if team_filter:
                    selected = {str(name).lower() for name in team_filter}
                    filtered_players = [p for p in filtered_players if (p.team or "").lower() in selected]
            entities: list[tuple[int, int | None]] = [(p.index, p.record_ptr) for p in filtered_players]
            snapshot = _build_player_snapshot(model, entities)
        elif entity_key == "teams":
            entities = [(idx, None) for idx, _ in model.team_list]
        elif entity_key == "staff":
            entities = [(idx, None) for idx, _ in model.staff_list]
        else:
            entities = [(idx, None) for idx, _ in model.stadium_list]
        progress_total = 0
        progress_current = 0
        if progress_cb is not None:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_lower = sheet_name.strip().lower()
                if entity_key == "teams" and sheet_lower == "team players":
                    continue
                header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
                if not header_row:
                    continue
                progress_total += len(entities)
            if progress_total > 0:
                progress_cb(0, progress_total, None)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_lower = sheet_name.strip().lower()
            if entity_key == "teams" and sheet_lower == "team players":
                continue
            header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
            if not header_row:
                continue
            headers = [cell.value for cell in header_row]
            preferred = _preferred_categories(sheet_name, categories)
            header_map, _ignored = _map_headers(model, headers, global_map, per_category, preferred)
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)
            if not entities:
                continue
            for row_idx, (entity_index, record_ptr) in enumerate(entities, start=2):
                record_view = None
                if snapshot is not None and entity_key == "players":
                    record_view = snapshot.record_view(int(entity_index))
                for col_idx, match in enumerate(header_map, start=1):
                    if match is None:
                        continue
                    try:
                        value = _decode_field_value_from_record(
                            model,
                            entity_type=entity_key[:-1] if entity_key.endswith("s") else entity_key,
                            category=match.category,
                            field_name=match.field_name,
                            meta=match.meta,
                            record=record_view,
                        )
                        if value is _FALLBACK:
                            value = model.decode_field_value(
                                entity_type=entity_key[:-1] if entity_key.endswith("s") else entity_key,
                                entity_index=int(entity_index),
                                category=match.category,
                                field_name=match.field_name,
                                meta=match.meta,
                                record_ptr=record_ptr,
                            )
                    except Exception:
                        value = None
                    cleaned = _sanitize_excel_value(value)
                    if isinstance(value, str) and cleaned != value:
                        sanitized_values += 1
                    sheet.cell(row=row_idx, column=col_idx, value=cleaned)
                if progress_cb is not None and progress_total > 0:
                    progress_current += 1
                    if progress_current % 5 == 0 or progress_current == progress_total:
                        progress_cb(progress_current, progress_total, sheet_name)
            result.rows_written = max(result.rows_written, len(entities))
            result.sheets_written += 1
        if sanitized_values:
            result.warnings.append(f"Sanitized {sanitized_values} string values with invalid Excel characters.")
        wb.save(output)
        return result


def import_players_from_excel(
    model: PlayerDataModel,
    workbook_path: str | Path,
    *,
    name_overrides: dict[str, str] | None = None,
    only_names: set[str] | None = None,
) -> ImportResult:
    return import_excel_workbook(
        model,
        workbook_path,
        "players",
        name_overrides=name_overrides,
        only_names=only_names,
    )


def import_teams_from_excel(
    model: PlayerDataModel,
    workbook_path: str | Path,
    *,
    name_overrides: dict[str, str] | None = None,
    only_names: set[str] | None = None,
) -> ImportResult:
    return import_excel_workbook(
        model,
        workbook_path,
        "teams",
        name_overrides=name_overrides,
        only_names=only_names,
    )


def import_staff_from_excel(
    model: PlayerDataModel,
    workbook_path: str | Path,
    *,
    name_overrides: dict[str, str] | None = None,
    only_names: set[str] | None = None,
) -> ImportResult:
    return import_excel_workbook(
        model,
        workbook_path,
        "staff",
        name_overrides=name_overrides,
        only_names=only_names,
    )


def import_stadiums_from_excel(
    model: PlayerDataModel,
    workbook_path: str | Path,
    *,
    name_overrides: dict[str, str] | None = None,
    only_names: set[str] | None = None,
) -> ImportResult:
    return import_excel_workbook(
        model,
        workbook_path,
        "stadiums",
        name_overrides=name_overrides,
        only_names=only_names,
    )


def export_players_to_excel(
    model: PlayerDataModel,
    output_path: str | Path,
    *,
    template_path: str | Path | None = None,
    team_filter: set[str] | None = None,
    players: Sequence[Player] | None = None,
) -> ExportResult:
    return export_excel_workbook(
        model,
        output_path,
        "players",
        template_path=template_path,
        team_filter=team_filter,
        players=players,
    )


def export_teams_to_excel(
    model: PlayerDataModel,
    output_path: str | Path,
    *,
    template_path: str | Path | None = None,
) -> ExportResult:
    return export_excel_workbook(model, output_path, "teams", template_path=template_path)


def export_staff_to_excel(
    model: PlayerDataModel,
    output_path: str | Path,
    *,
    template_path: str | Path | None = None,
) -> ExportResult:
    return export_excel_workbook(model, output_path, "staff", template_path=template_path)


def export_stadiums_to_excel(
    model: PlayerDataModel,
    output_path: str | Path,
    *,
    template_path: str | Path | None = None,
) -> ExportResult:
    return export_excel_workbook(model, output_path, "stadiums", template_path=template_path)


__all__ = [
    "ExportResult",
    "ImportResult",
    "export_excel_workbook",
    "export_players_to_excel",
    "export_staff_to_excel",
    "export_stadiums_to_excel",
    "export_teams_to_excel",
    "import_excel_workbook",
    "import_players_from_excel",
    "import_staff_from_excel",
    "import_stadiums_from_excel",
    "import_teams_from_excel",
    "template_path_for",
]